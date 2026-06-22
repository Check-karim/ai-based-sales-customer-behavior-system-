import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from services.ai_analytics import (
  get_sales_summary,
  get_monthly_sales,
  get_top_products,
  segment_customers_rfm,
  predict_churn_risk,
  get_purchase_patterns,
  get_payment_method_distribution,
  get_loyalty_tier_analysis,
  generate_behavior_insights,
  get_sales_forecast,
)


def _header_style():
  return {
    "font": Font(bold=True, color="FFFFFF", size=11),
    "fill": PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid"),
    "alignment": Alignment(horizontal="center", vertical="center"),
  }


def _apply_header(ws, row, cols):
  style = _header_style()
  for col in range(1, cols + 1):
    cell = ws.cell(row=row, column=col)
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]


def generate_excel_report():
  wb = Workbook()

  summary = get_sales_summary()
  ws_summary = wb.active
  ws_summary.title = "Sales Summary"
  ws_summary.append(["Deluxe Supermarket - Sales Report"])
  ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
  ws_summary.append([])
  ws_summary.append(["Metric", "Value"])
  _apply_header(ws_summary, 4, 2)
  ws_summary.append(["Total Transactions", summary["total_transactions"]])
  ws_summary.append(["Total Revenue", f"${float(summary['total_revenue']):,.2f}"])
  ws_summary.append(["Average Transaction", f"${float(summary['avg_transaction']):,.2f}"])
  ws_summary.append(["Unique Customers", summary["unique_customers"]])
  ws_summary.column_dimensions["A"].width = 25
  ws_summary.column_dimensions["B"].width = 20

  ws_monthly = wb.create_sheet("Monthly Sales")
  ws_monthly.append(["Month", "Transactions", "Revenue"])
  _apply_header(ws_monthly, 1, 3)
  for row in get_monthly_sales():
    ws_monthly.append([row["month"], row["transactions"], float(row["revenue"])])
  for col in ["A", "B", "C"]:
    ws_monthly.column_dimensions[col].width = 18

  ws_products = wb.create_sheet("Top Products")
  ws_products.append(["Product", "Category", "Units Sold", "Revenue"])
  _apply_header(ws_products, 1, 4)
  for row in get_top_products():
    ws_products.append([
      row["product_name"],
      row["category"],
      row["units_sold"],
      float(row["revenue"]),
    ])
  for col in ["A", "B", "C", "D"]:
    ws_products.column_dimensions[col].width = 20

  ws_segments = wb.create_sheet("Customer Segments")
  ws_segments.append(["Customer", "Loyalty Tier", "Recency (days)", "Frequency", "Monetary", "AI Segment", "AI Score"])
  _apply_header(ws_segments, 1, 7)
  for row in segment_customers_rfm():
    ws_segments.append([
      row["customer_name"],
      row["loyalty_tier"],
      row["recency_days"] or "N/A",
      row["frequency"],
      float(row["monetary"]),
      row.get("segment", "N/A"),
      row.get("ai_score", 0),
    ])
  for col in ["A", "B", "C", "D", "E", "F", "G"]:
    ws_segments.column_dimensions[col].width = 18

  ws_churn = wb.create_sheet("Churn Risk")
  ws_churn.append(["Customer", "Loyalty Tier", "Recency", "Frequency", "Monetary", "Churn %", "Risk Level"])
  _apply_header(ws_churn, 1, 7)
  for row in predict_churn_risk():
    ws_churn.append([
      row["customer_name"],
      row["loyalty_tier"],
      row["recency_days"],
      row["frequency"],
      row["monetary"],
      row["churn_probability"],
      row["risk_level"],
    ])
  for col in ["A", "B", "C", "D", "E", "F", "G"]:
    ws_churn.column_dimensions[col].width = 18

  ws_patterns = wb.create_sheet("Purchase Patterns")
  ws_patterns.append(["Category", "Purchase Count", "Revenue", "Avg Quantity"])
  _apply_header(ws_patterns, 1, 4)
  for row in get_purchase_patterns():
    ws_patterns.append([
      row["category"],
      row["purchase_count"],
      float(row["category_revenue"]),
      round(float(row["avg_quantity"]), 2),
    ])
  for col in ["A", "B", "C", "D"]:
    ws_patterns.column_dimensions[col].width = 20

  ws_insights = wb.create_sheet("AI Insights")
  ws_insights.append(["Customer", "Insight Type", "Value", "Confidence %"])
  _apply_header(ws_insights, 1, 4)
  for row in generate_behavior_insights():
    ws_insights.append([
      row["customer_name"],
      row["type"],
      row["value"],
      row["confidence"],
    ])
  for col in ["A", "B", "C", "D"]:
    ws_insights.column_dimensions[col].width = 25

  buffer = io.BytesIO()
  wb.save(buffer)
  buffer.seek(0)
  return buffer


def _pdf_table(data, col_widths=None):
  table = Table(data, colWidths=col_widths)
  table.setStyle(TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4D3E")),
    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ("FONTSIZE", (0, 0), (-1, -1), 9),
    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F7F4")]),
    ("TOPPADDING", (0, 0), (-1, -1), 6),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
  ]))
  return table


def generate_pdf_report():
  buffer = io.BytesIO()
  doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5 * inch)
  styles = getSampleStyleSheet()
  title_style = ParagraphStyle(
    "CustomTitle",
    parent=styles["Heading1"],
    fontSize=18,
    textColor=colors.HexColor("#1B4D3E"),
    spaceAfter=12,
  )
  subtitle_style = ParagraphStyle(
    "CustomSubtitle",
    parent=styles["Normal"],
    fontSize=10,
    textColor=colors.grey,
    spaceAfter=20,
  )
  section_style = ParagraphStyle(
    "Section",
    parent=styles["Heading2"],
    fontSize=13,
    textColor=colors.HexColor("#1B4D3E"),
    spaceBefore=16,
    spaceAfter=8,
  )

  elements = []
  elements.append(Paragraph("Deluxe Supermarket", title_style))
  elements.append(Paragraph("AI Sales & Customer Behavior Report", subtitle_style))
  elements.append(Paragraph(
    f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}",
    subtitle_style,
  ))

  summary = get_sales_summary()
  elements.append(Paragraph("Sales Summary", section_style))
  summary_data = [
    ["Metric", "Value"],
    ["Total Transactions", str(summary["total_transactions"])],
    ["Total Revenue", f"${float(summary['total_revenue']):,.2f}"],
    ["Average Transaction", f"${float(summary['avg_transaction']):,.2f}"],
    ["Unique Customers", str(summary["unique_customers"])],
  ]
  elements.append(_pdf_table(summary_data, [3 * inch, 3 * inch]))
  elements.append(Spacer(1, 12))

  forecast = get_sales_forecast()
  elements.append(Paragraph("AI Sales Forecast", section_style))
  forecast_data = [
    ["Forecast Revenue", "Growth Rate", "Confidence"],
    [
      f"${forecast['forecast_revenue']:,.2f}",
      f"{forecast['growth_rate']}%",
      f"{forecast['confidence']}%",
    ],
  ]
  elements.append(_pdf_table(forecast_data, [2.5 * inch, 2.5 * inch, 2.5 * inch]))
  elements.append(Spacer(1, 12))

  elements.append(Paragraph("Top Products", section_style))
  products_data = [["Product", "Category", "Units Sold", "Revenue"]]
  for row in get_top_products(8):
    products_data.append([
      row["product_name"][:25],
      row["category"],
      str(row["units_sold"]),
      f"${float(row['revenue']):,.2f}",
    ])
  elements.append(_pdf_table(products_data, [2.5 * inch, 1.8 * inch, 1.2 * inch, 1.5 * inch]))
  elements.append(Spacer(1, 12))

  elements.append(Paragraph("Customer Segments (AI)", section_style))
  segments_data = [["Customer", "Tier", "Segment", "AI Score"]]
  for row in segment_customers_rfm()[:10]:
    segments_data.append([
      row["customer_name"],
      row["loyalty_tier"],
      row.get("segment", "N/A"),
      str(row.get("ai_score", 0)),
    ])
  elements.append(_pdf_table(segments_data, [2.5 * inch, 1.5 * inch, 2 * inch, 1.5 * inch]))
  elements.append(Spacer(1, 12))

  elements.append(Paragraph("Churn Risk Analysis", section_style))
  churn_data = [["Customer", "Churn %", "Risk Level", "Monetary"]]
  for row in predict_churn_risk()[:8]:
    churn_data.append([
      row["customer_name"],
      f"{row['churn_probability']}%",
      row["risk_level"],
      f"${row['monetary']:,.2f}",
    ])
  elements.append(_pdf_table(churn_data, [2.5 * inch, 1.5 * inch, 1.5 * inch, 1.8 * inch]))
  elements.append(Spacer(1, 12))

  elements.append(Paragraph("AI Behavior Insights", section_style))
  insights_data = [["Customer", "Type", "Insight", "Confidence"]]
  for row in generate_behavior_insights()[:10]:
    insights_data.append([
      row["customer_name"][:20],
      row["type"],
      str(row["value"])[:35],
      f"{row['confidence']}%",
    ])
  elements.append(_pdf_table(insights_data, [2 * inch, 1.2 * inch, 3 * inch, 1.2 * inch]))

  doc.build(elements)
  buffer.seek(0)
  return buffer
